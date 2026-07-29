#!/usr/bin/env python3
"""Task 4 — verify every figure in chapter 2 of docs/slide_plan_10-16.md.

Corpus of record (task 4 instruction): only
  data/output/2511-2604_hybrid.csv
  data/output/sentiment_classified_hybrid.csv
Anything whose source lies outside those two is labelled with its own source
so the reader can see what was checked against what.

Writes docs/slide_numbers_check.md. The spec itself is never modified.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import statistics
from collections import Counter, defaultdict
from pathlib import Path

from filter_ads_202511_202604 import keyword_matches
from normalize_gold_standard_192 import (
    DEFAULT_OUTPUT as NORMALIZED_GOLD_STANDARD,
    validate_normalized_gold,
)
from slide_number_definitions import (
    EXCHANGE_ACCOUNT_KEY,
    EXCHANGE_CATEGORY,
    has_exchange_ratio,
    has_honorific_consideration,
    is_exchange_template,
    is_platform_reply,
    top_fraction_account_count,
    wald_interval,
)

csv.field_size_limit(10**9)

HYBRID = Path("data/output/sentiment_classified_hybrid.csv")
HYBRID_CORPUS = Path("data/output/2511-2604_hybrid.csv")
FINAL = Path("data/output/2511-2604_final.csv")
OLD = Path("data/output/2511-2604.csv")
ADDL = Path("data/output/removed_additional_ads_with_reasons_202511_202604.csv")
GOLD192 = NORMALIZED_GOLD_STANDARD
XLSX = Path("outputs/sentiment-analysis-20260716/category_random_sample_30_each.xlsx")
NEG1 = Path("outputs/negotiation-unclassified-20260728/negotiation_expressions_not_exchange_random50.csv")
NEG2 = Path("data/output/random_sample_50_negotiation_not_exchange_202511_202604.csv")
MONTHS = ["202511", "202512", "202601", "202602", "202603", "202604"]
OUT = Path("docs/slide_numbers_check.md")
SPEC = Path("docs/slide_plan_10-16.md")

CATEGORIES = ["不満・怒り", "焦り・競争", "交換・取引", "欲望・執着",
              "喜び・満足", "情報共有", "中立"]
COLMAP = {"喜び満足": "喜び・満足", "欲望執着": "欲望・執着", "不満怒り": "不満・怒り",
          "焦り競争": "焦り・競争", "情報共有": "情報共有", "交換取引": "交換・取引"}

results: list[tuple] = []


def check(section, item, spec, actual, source, tol=0.0):
    """Record one comparison. spec/actual may be numbers or strings."""
    if isinstance(spec, (int, float)) and isinstance(actual, (int, float)):
        ok = abs(float(spec) - float(actual)) <= tol
    else:
        ok = str(spec).strip() == str(actual).strip()
    results.append((section, item, spec, actual, "一致" if ok else "★不一致", source))
    return ok


def sha16(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def ids(p: Path, col="投稿ID_文字列") -> set[str]:
    with p.open("r", encoding="utf-8-sig", newline="") as f:
        return {r[col] for r in csv.DictReader(f)}


def main(
    output: Path = OUT,
    gold192: Path = GOLD192,
) -> None:
    results.clear()
    gold192 = validate_normalized_gold(gold192)["path"]
    gold192 = Path(gold192)
    # ---------- 2-0 ----------
    check("2-0", "SHA-256 コーパス", "3bf78817892b356b", sha16(HYBRID_CORPUS), "実ファイル")
    check("2-0", "SHA-256 分類結果", "f273c9306507804a", sha16(HYBRID), "実ファイル")

    allrows: dict[str, list[str]] = {}
    for m in MONTHS:
        with Path(f"{m}.csv").open("r", encoding="utf-8-sig", newline="") as f:
            r = csv.reader(f)
            next(r)
            for row in r:
                allrows[row[1]] = row
    allid = set(allrows)
    hyb, fin, old, addl = ids(HYBRID), ids(FINAL), ids(OLD), ids(ADDL)

    s3 = old - fin
    s1 = (allid - fin) - addl - s3
    s2 = addl
    removed = allid - hyb

    check("2-0", "月別原票（投稿ID一意）", 136288, len(allid), "月別CSV6件")
    check("2-0", "① 最終版キーワード除去", 19362, len(s1), "集合演算")
    check("2-0", "② 旧版の追加広告分類", 5874, len(s2), ADDL.name)
    check("2-0", "③ 最終版が新たに除去", 134, len(s3), "old − final")
    check("2-0", "①後の残数", 116926, len(allid) - len(s1), "計算")
    check("2-0", "②後の残数", 111052, len(allid) - len(s1) - len(s2), "計算")
    check("2-0", "分析対象", 110918, len(hyb), HYBRID.name)
    check("2-0", "除去合計", 25370, len(removed), "計算")
    check("2-0", "除去率(%)", 18.61, round(len(removed) / len(allid) * 100, 2), "計算", 0.005)
    check("2-0", "3集合が相互排他", "重複0",
          "重複0" if not (s1 & s2 or s1 & s3 or s2 & s3) else "重複あり", "集合演算")
    check("2-0", "①②③の合計＝除去合計", len(removed), len(s1) + len(s2) + len(s3), "計算")

    restored = fin - old  # 旧版が除去し最終版が残した投稿
    check("2-0", "復帰分（旧版除去→最終版で復活）", 5615, len(restored), "final − old")
    dai = sum(1 for i in restored if keyword_matches(allrows[i][8]) == ["第弾"])
    check("2-0", "`第弾` の過剰除去", 2002, dai,
          "final − old のうち旧フィルタ一致が `第弾` のみ")
    check("2-0", "最終版が取りこぼした広告", 3600, len(addl & fin), "addl ∩ final")

    before = Counter()
    for i in allid:
        pass  # counted below from the pre-removal classification
    # pre-removal classification: hybrid rows + removed rows classified with v2
    from classify_sns_rule_based import classify_detailed
    hyb_cat = {}
    with HYBRID.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            hyb_cat[row["投稿ID_文字列"]] = row["sentiment_category"]
    after = Counter(hyb_cat.values())
    before = Counter(after)
    for i in removed:
        before[classify_detailed(allrows[i][8]).primary] += 1

    spec_tbl = {"情報共有": (12392, 4106, 66.87), "中立": (58294, 45418, 22.09),
                "焦り・競争": (4148, 3272, 21.12), "喜び・満足": (15131, 13002, 14.07),
                "欲望・執着": (10185, 9513, 6.60), "不満・怒り": (11586, 11291, 2.55),
                "交換・取引": (24552, 24316, 0.96)}
    for cat, (b, a, rate) in spec_tbl.items():
        check("2-0", f"除去前 {cat}", b, before[cat], "136,288件をv2で分類")
        check("2-0", f"ハイブリッド {cat}", a, after[cat], HYBRID.name)
        check("2-0", f"除去率 {cat}(%)", rate, round((before[cat] - after[cat]) / before[cat] * 100, 2),
              "計算", 0.005)
    check("2-0", "スライド3 収集(万件)", 13.6, round(len(allid) / 10000, 1), "計算", 0.05)
    check("2-0", "スライド3 広告(万件)", 2.5, round(len(removed) / 10000, 1), "計算", 0.05)
    check("2-0", "スライド3 分析対象(万件)", 11.1, round(len(hyb) / 10000, 1), "計算", 0.05)

    # ---------- 2-1 ----------
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        ws = wb["ランダム抽出"]
        hdr = [c.value for c in ws[1]]
        ci = hdr.index("内容") + 1
        cat_i = next(i for i, h in enumerate(hdr, 1)
                     if h and ("sentiment" in str(h) or "カテゴリ" in str(h)))
        buckets: dict[str, Counter] = defaultdict(Counter)
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, ci)
            cat = ws.cell(r, cat_i).value
            if not cat:
                continue
            rgb = getattr(cell.fill.fgColor, "rgb", None) or ""
            rgb = str(rgb)[-6:].upper()
            hexes = {"00B050": "緑", "92D050": "緑", "C6EFCE": "緑",
                     "FFFF00": "黄", "FFEB9C": "黄", "FFC000": "黄",
                     "FF0000": "赤", "FFC7CE": "赤", "FF9999": "赤"}
            buckets[str(cat)][hexes.get(rgb, "無")] += 1
        spec21 = {"焦り・競争": (30, 24, 4, 2, 80), "喜び・満足": (30, 14, 15, 1, 47),
                  "情報共有": (23, 10, 1, 12, 43)}
        for cat, (n, g, y, r, rate) in spec21.items():
            b = buckets.get(cat, Counter())
            got_n = b["緑"] + b["黄"] + b["赤"]
            check("2-1", f"{cat} 検討数", n, got_n, XLSX.name)
            check("2-1", f"{cat} 緑", g, b["緑"], XLSX.name)
            check("2-1", f"{cat} 黄", y, b["黄"], XLSX.name)
            check("2-1", f"{cat} 赤", r, b["赤"], XLSX.name)
            if got_n:
                check("2-1", f"{cat} 一致率(%)", rate, round(b["緑"] / got_n * 100), XLSX.name, 0.6)
    except Exception as e:  # noqa: BLE001
        results.append(("2-1", "セル塗りの読み取り", "—", f"検証不可: {type(e).__name__}",
                        "★検証不可", XLSX.name))

    # ---------- 2-3 ----------
    with gold192.open("r", encoding="utf-8-sig", newline="") as f:
        gold = [r for r in csv.DictReader(f) if r["要検討"] != "1"]
    n = len(gold)
    check("2-3", "正解セット件数", 192, n, gold192.name)
    gc = {c: sum(1 for r in gold if any(r[col] == "1" for col, cc in COLMAP.items() if cc == c))
          for c in CATEGORIES if c != "中立"}
    gc["中立"] = sum(1 for r in gold if r["中立"] == "1")
    spec23 = {"交換・取引": (28.1, "21.8 – 34.5%"), "中立": (25.0, "18.9 – 31.1%"),
              "欲望・執着": (18.8, "13.2 – 24.3%"), "喜び・満足": (17.2, "11.9 – 22.5%"),
              "焦り・競争": (14.1, "9.1 – 19.0%"), "不満・怒り": (7.3, "3.6 – 11.0%"),
              "情報共有": (2.6, "0.4 – 4.9%")}
    for cat, (pct, ci_s) in spec23.items():
        check("2-3", f"{cat} 割合(%)", pct, round(gc[cat] / n * 100, 1),
              gold192.name, 0.05)
        lo, hi = wald_interval(gc[cat], n)
        check("2-3", f"{cat} 95%CI", ci_s, f"{lo * 100:.1f} – {hi * 100:.1f}%",
              "Wald正規近似（95%, z=1.96, n=192, カテゴリ別二値）")
    check("2-3", "ハイブリッド追加分", 2015, len(hyb - old), "hybrid − old")

    # ---------- 2-4 ----------
    ex = []
    with HYBRID.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            if row["sentiment_category"] == EXCHANGE_CATEGORY:
                ex.append(row)
    check("2-4", "交換・取引 投稿数", 24316, len(ex), HYBRID.name)
    check("2-4", "一意 ユーザーID", 10677,
          len({r[EXCHANGE_ACCOUNT_KEY].strip() for r in ex}), HYBRID.name)
    check("2-4", "一意 アカウントID", 10715, len({r["アカウントID"].strip() for r in ex}), HYBRID.name)
    check("2-4", "一意 (アカウントID,名前)",
          10894, len({(r["アカウントID"].strip(), r["名前"].strip()) for r in ex}), HYBRID.name)
    check("2-4", "一意 名前", 8907, len({r["名前"].strip() for r in ex}), HYBRID.name)
    by = defaultdict(list)
    for r in ex:
        by[r[EXCHANGE_ACCOUNT_KEY].strip()].append(r)
    counts = sorted((len(v) for v in by.values()), reverse=True)
    n_acc = len(by)
    check("2-4", "平均投稿数", 2.28, round(len(ex) / n_acc, 2), "計算", 0.005)
    check("2-4", "中央値", 1, statistics.median(counts), "計算")
    singles = sum(1 for c in counts if c == 1)
    check("2-4", "1件のみのアカウント", 7198, singles, "計算")
    check("2-4", "1件のみの割合(%)", 67.4, round(singles / n_acc * 100, 1), "計算", 0.05)
    check("2-4", "上位30の合計", 1796, sum(counts[:30]), "計算")
    check("2-4", "上位30シェア(%)", 7.4, round(sum(counts[:30]) / len(ex) * 100, 1), "計算", 0.05)
    k1 = top_fraction_account_count(n_acc, 0.01)
    k10 = top_fraction_account_count(n_acc, 0.10)
    check("2-4", "上位1%のアカウント数", 106, k1, "floor(10,677 × 0.01)")
    check("2-4", "上位1%シェア(%)", 14.9,
          round(sum(counts[:k1]) / len(ex) * 100, 1), "3,619 / 24,316", 0.05)
    check("2-4", "上位10%シェア(%)", 45.6,
          round(sum(counts[:k10]) / len(ex) * 100, 1), "11,082 / 24,316", 0.05)
    t = sum(is_exchange_template(r["内容"]) for r in ex)
    check("2-4", "定型書式 件数", 12411, t, "共通の定型書式定義")
    check("2-4", "定型書式 割合(%)", 51.0, round(t / len(ex) * 100, 1),
          "12,411 / 24,316", 0.05)

    # ---------- 2-5 ----------
    try:
        neg = {}
        for p in (NEG1, NEG2):
            with p.open("r", encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    key = row.get("投稿ID_文字列") or row.get("post_id") or row.get("投稿ID")
                    neg[key] = row
        check("2-5", "実質件数", 98, len(neg), f"{NEG1.name}+{NEG2.name}")
        txts = {k: (v.get("内容") or v.get("clean_text") or "") for k, v in neg.items()}
        reply = sum(is_platform_reply(v) for v in neg.values())
        check("2-5", "リプライ件数", 79, reply, "リプライ先の投稿IDが非空")
        cat = {k: hyb_cat.get(k) for k in neg}
        check("2-5", "中立と分類されていた件数", 90,
              sum(1 for v in cat.values() if v == "中立"), HYBRID.name)
        string_checks = [
            ("「検索より/から失礼いたします」", 48,
             lambda text: re.search(r"検索(?:より|から)", text) is not None,
             r"検索(?:より|から)"),
            ("「ご検討」", 42, has_honorific_consideration, r"(?:ご\|御)検討"),
            ("「初めまして」", 31,
             lambda text: re.search(r"初めまして|はじめまして", text) is not None,
             r"初めまして\|はじめまして"),
            ("🙇", 31, lambda text: "🙇" in text, "本文に 🙇"),
            ("交換比率(n:m)", 17, has_exchange_ratio, r"\d\s*[:：]\s*\d"),
            ("「比率違い」", 3, lambda text: "比率違い" in text, "本文に 比率違い"),
        ]
        for label, spec_n, predicate, source in string_checks:
            check("2-5", label, spec_n, sum(predicate(text) for text in txts.values()), source)
    except Exception as e:  # noqa: BLE001
        results.append(("2-5", "交渉表現98件", "—", f"検証不可: {type(e).__name__}",
                        "★検証不可", "—"))

    # ---------- output ----------
    ok = sum(1 for r in results if r[4] == "一致")
    ng = sum(1 for r in results if r[4] == "★不一致")
    na = sum(1 for r in results if r[4] == "★検証不可")
    L = ["# 仕様書 第2章 数値照合結果（タスク4）", "",
         "**照合日**: 2026-07-30",
         f"**照合対象仕様書**: `{SPEC}`（SHA-256 `{sha16(SPEC)}`）",
         "**コーパス**: `data/output/2511-2604_hybrid.csv` / "
         "`data/output/sentiment_classified_hybrid.csv`",
         f"**正解セット**: `{gold192.name}`", "",
         f"## 集計: 一致 {ok} / 不一致 {ng} / 検証不可 {na}（全 {len(results)} 項目）", "",
         "| 節 | 項目 | 仕様書 | 実測 | 判定 | 照合元 |",
         "|---|---|---:|---:|---|---|"]
    for sec, item, spec, act, verdict, src in results:
        L.append(f"| {sec} | {item} | {spec} | {act} | {verdict} | {src} |")
    L.append("")
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("\n".join(L) + "\n", encoding="utf-8")

    print(f"一致 {ok} / 不一致 {ng} / 検証不可 {na}  (全 {len(results)})")
    for r in results:
        if r[4] != "一致":
            print(f"  [{r[0]}] {r[1]}: 仕様書={r[2]} 実測={r[3]}  ({r[4]})")
    print(f"-> {output}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=OUT)
    parser.add_argument("--gold", type=Path, default=GOLD192)
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    main(args.output, args.gold)
